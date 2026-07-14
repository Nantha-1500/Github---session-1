import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def create_test_summary_report():
    """
    Create a Test Summary Report based on the template format
    and store it in the outputs folder
    
    Test Summary Report - Base MSME
    Total Testcases: 259
    Total Testcases Executed: 259
    Test cases passed: 246
    Test cases failed: 13
    """
    
    # Load the template
    template_path = "templates/MSME-Base-Test-Cases-V1.xlsx"
    wb = openpyxl.load_workbook(template_path)
    
    # Create a new workbook for the summary report
    summary_wb = openpyxl.Workbook()
    summary_ws = summary_wb.active
    summary_ws.title = "Test Summary"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    summary_header_font = Font(bold=True, size=11, color="FFFFFF")
    summary_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add report title
    summary_ws['A1'] = "TEST SUMMARY REPORT - BASE MSME"
    summary_ws['A1'].font = title_font
    summary_ws['A1'].fill = title_fill
    summary_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    summary_ws.merge_cells('A1:F1')
    summary_ws.row_dimensions[1].height = 25
    
    # Add report metadata
    row = 3
    summary_ws[f'A{row}'] = "Report Generated:"
    summary_ws[f'A{row}'].font = Font(bold=True)
    summary_ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 1
    
    summary_ws[f'A{row}'] = "Template Used:"
    summary_ws[f'A{row}'].font = Font(bold=True)
    summary_ws[f'B{row}'] = "MSME-Base-Test-Cases-V1.xlsx"
    row += 2
    
    # Add SUMMARY STATISTICS section
    summary_ws[f'A{row}'] = "SUMMARY STATISTICS"
    summary_ws[f'A{row}'].font = summary_header_font
    summary_ws[f'A{row}'].fill = summary_header_fill
    summary_ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    summary_ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Define statistics data
    stats_data = [
        ["Total Testcases", 259],
        ["Total Testcases Executed", 259],
        ["Test cases passed", 246],
        ["Test cases failed", 13],
    ]
    
    # Add statistics with styling
    for stat_label, stat_value in stats_data:
        summary_ws[f'A{row}'] = stat_label
        summary_ws[f'A{row}'].font = Font(bold=True)
        summary_ws[f'A{row}'].border = border
        summary_ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        summary_ws[f'B{row}'] = stat_value
        summary_ws[f'B{row}'].font = Font(size=11)
        summary_ws[f'B{row}'].border = border
        summary_ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    row += 1
    
    # Calculate pass rate
    total_tests = 259
    passed_tests = 246
    failed_tests = 13
    pass_rate = (passed_tests / total_tests) * 100
    
    # Add Pass Rate section
    summary_ws[f'A{row}'] = "Pass Rate"
    summary_ws[f'A{row}'].font = Font(bold=True)
    summary_ws[f'A{row}'].border = border
    
    summary_ws[f'B{row}'] = f"{pass_rate:.2f}%"
    summary_ws[f'B{row}'].font = Font(size=11, bold=True)
    summary_ws[f'B{row}'].border = border
    summary_ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Add execution summary
    summary_ws[f'A{row}'] = "Execution Summary"
    summary_ws[f'A{row}'].font = Font(bold=True, size=11)
    summary_ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    summary_ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Add execution details
    summary_ws[f'A{row}'] = "Status:"
    summary_ws[f'A{row}'].font = Font(bold=True)
    if failed_tests == 0:
        summary_ws[f'B{row}'] = "✓ ALL TESTS PASSED"
        summary_ws[f'B{row}'].font = Font(color="00B050", bold=True, size=11)
    else:
        summary_ws[f'B{row}'] = f"✗ {failed_tests} TEST(S) FAILED"
        summary_ws[f'B{row}'].font = Font(color="FF0000", bold=True, size=11)
    row += 1
    
    # Add column headers for detailed test breakdown
    row += 1
    headers = ["Test Result", "Count", "Percentage"]
    for col_idx, header in enumerate(headers, 1):
        cell = summary_ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Add test breakdown
    breakdown_data = [
        ["Passed", passed_tests, f"{(passed_tests/total_tests)*100:.2f}%"],
        ["Failed", failed_tests, f"{(failed_tests/total_tests)*100:.2f}%"],
        ["Total", total_tests, "100.00%"],
    ]
    
    for data_row in breakdown_data:
        for col_idx, value in enumerate(data_row, 1):
            cell = summary_ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx == 1:
                cell.font = Font(bold=True)
        row += 1
    
    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 20
    summary_ws.column_dimensions['C'].width = 18
    summary_ws.column_dimensions['D'].width = 15
    summary_ws.column_dimensions['E'].width = 15
    summary_ws.column_dimensions['F'].width = 20
    
    # Create outputs directory if it doesn't exist
    os.makedirs("outputs", exist_ok=True)
    
    # Save the report
    output_filename = f"outputs/Test_Summary_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    summary_wb.save(output_filename)
    
    print("=" * 60)
    print("✓ TEST SUMMARY REPORT CREATED SUCCESSFULLY")
    print("=" * 60)
    print(f"Report Title: Test Summary Report - Base MSME")
    print(f"Total Testcases: {total_tests}")
    print(f"Total Testcases Executed: {total_tests}")
    print(f"Test cases passed: {passed_tests}")
    print(f"Test cases failed: {failed_tests}")
    print(f"Pass Rate: {pass_rate:.2f}%")
    print("=" * 60)
    print(f"✓ File saved to: {output_filename}")
    print("=" * 60)
    return output_filename

if __name__ == "__main__":
    create_test_summary_report()
