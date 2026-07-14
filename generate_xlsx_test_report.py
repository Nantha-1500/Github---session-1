import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def analyze_excel_and_generate_xlsx_report(excel_file):
    """
    Analyze Excel file and generate test summary report in XLSX format
    """
    
    try:
        # Read Excel file
        print("\n" + "="*80)
        print("TEST SUMMARY REPORT GENERATOR")
        print("="*80 + "\n")
        
        excel_file_obj = pd.ExcelFile(excel_file)
        
        print(f"📁 File: {excel_file}")
        print(f"📊 Sheets: {excel_file_obj.sheet_names}\n")
        
        # Initialize counters
        total_test_cases = 0
        total_executed = 0
        total_passed = 0
        total_failed = 0
        
        all_sheets_data = {}
        
        # Analyze each sheet
        for sheet in excel_file_obj.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet)
            
            print(f"\n{'─'*80}")
            print(f"📋 Sheet: {sheet}")
            print(f"{'─'*80}")
            print(f"Rows: {len(df)}, Columns: {len(df.columns)}")
            print(f"Columns: {list(df.columns)}\n")
            
            # Display first few rows
            print(df.head(10).to_string())
            print()
            
            all_sheets_data[sheet] = df
            
            # Find status column
            status_col = None
            for col in df.columns:
                col_lower = col.lower()
                if any(keyword in col_lower for keyword in ['status', 'result', 'outcome', 'test status', 'execution status']):
                    status_col = col
                    break
            
            # Count test cases
            sheet_total = len(df)
            total_test_cases += sheet_total
            
            # Count by status
            if status_col and df[status_col].notna().any():
                print(f"✓ Found Status Column: '{status_col}'\n")
                
                status_counts = df[status_col].value_counts()
                print("Status Distribution:")
                print(status_counts)
                print()
                
                sheet_passed = 0
                sheet_failed = 0
                sheet_executed = 0
                
                for status, count in status_counts.items():
                    status_str = str(status).lower().strip()
                    
                    if any(keyword in status_str for keyword in ['pass', 'passed', 'success', 'ok']):
                        sheet_passed += count
                    elif any(keyword in status_str for keyword in ['fail', 'failed', 'error', 'fault']):
                        sheet_failed += count
                    
                    sheet_executed += count
                
                total_executed += sheet_executed
                total_passed += sheet_passed
                total_failed += sheet_failed
        
        # Calculate rates
        pass_rate = (total_passed / total_executed * 100) if total_executed > 0 else 0
        fail_rate = (total_failed / total_executed * 100) if total_executed > 0 else 0
        execution_rate = (total_executed / total_test_cases * 100) if total_test_cases > 0 else 0
        
        # Create XLSX Report
        print("\n" + "="*80)
        print("CREATING EXCEL SUMMARY REPORT")
        print("="*80 + "\n")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        metric_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        metric_header_font = Font(bold=True, color="000000", size=11)
        
        metric_value_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        metric_value_font = Font(bold=True, size=12, color="000000")
        
        passed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        passed_font = Font(bold=True, color="006100", size=12)
        
        failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        failed_font = Font(bold=True, color="9C0006", size=12)
        
        execution_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        execution_font = Font(bold=True, color="375623", size=12)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:C{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "TEST SUMMARY REPORT"
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = center_alignment
        ws.row_dimensions[row].height = 30
        row += 1
        
        # Date
        ws.merge_cells(f'A{row}:C{row}')
        date_cell = ws[f'A{row}']
        date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = center_alignment
        ws.row_dimensions[row].height = 18
        row += 2
        
        # TEST EXECUTION METRICS Section
        ws.merge_cells(f'A{row}:C{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "📊 TEST EXECUTION METRICS"
        section_cell.font = Font(bold=True, size=12, color="FFFFFF")
        section_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        section_cell.alignment = left_alignment
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Metrics data
        metrics_data = [
            ("Total Test Cases", total_test_cases, None),
            ("Total Executed Tests", total_executed, None),
            ("Passed Tests ✅", total_passed, passed_fill),
            ("Failed Tests ❌", total_failed, failed_fill),
            ("Skipped Tests ⏭️", total_test_cases - total_executed, None),
        ]
        
        for metric_name, metric_value, cell_fill in metrics_data:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = metric_header_font
            ws[f'A{row}'].fill = metric_header_fill
            ws[f'A{row}'].border = border
            ws[f'A{row}'].alignment = left_alignment
            
            ws[f'B{row}'] = metric_value
            if cell_fill:
                ws[f'B{row}'].fill = cell_fill
                if metric_name == "Passed Tests ✅":
                    ws[f'B{row}'].font = passed_font
                elif metric_name == "Failed Tests ❌":
                    ws[f'B{row}'].font = failed_font
            else:
                ws[f'B{row}'].font = metric_value_font
                ws[f'B{row}'].fill = metric_value_fill
            
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = center_alignment
            
            ws.row_dimensions[row].height = 25
            row += 1
        
        row += 1
        
        # PASS/FAIL ANALYSIS Section
        ws.merge_cells(f'A{row}:C{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "📈 PASS/FAIL ANALYSIS"
        section_cell.font = Font(bold=True, size=12, color="FFFFFF")
        section_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        section_cell.alignment = left_alignment
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Analysis metrics
        analysis_data = [
            ("Pass Rate %", f"{pass_rate:.2f}%", passed_fill, passed_font),
            ("Fail Rate %", f"{fail_rate:.2f}%", failed_fill, failed_font),
            ("Execution Rate %", f"{execution_rate:.2f}%", execution_fill, execution_font),
        ]
        
        for metric_name, metric_value, cell_fill, cell_font in analysis_data:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = metric_header_font
            ws[f'A{row}'].fill = metric_header_fill
            ws[f'A{row}'].border = border
            ws[f'A{row}'].alignment = left_alignment
            
            ws[f'B{row}'] = metric_value
            ws[f'B{row}'].fill = cell_fill
            ws[f'B{row}'].font = cell_font
            ws[f'B{row}'].border = border
            ws[f'B{row}'].alignment = center_alignment
            
            ws.row_dimensions[row].height = 25
            row += 1
        
        row += 1
        
        # SUMMARY Section
        ws.merge_cells(f'A{row}:C{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "📋 SUMMARY"
        section_cell.font = Font(bold=True, size=12, color="FFFFFF")
        section_cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        section_cell.alignment = left_alignment
        ws.row_dimensions[row].height = 25
        row += 1
        
        summary_text = [
            f"✓ Out of {total_test_cases} total test cases, {total_executed} were executed.",
            f"✅ {total_passed} tests PASSED ({pass_rate:.2f}%)",
            f"❌ {total_failed} tests FAILED ({fail_rate:.2f}%)",
            f"📊 Execution Coverage: {execution_rate:.2f}%",
        ]
        
        for summary_line in summary_text:
            ws.merge_cells(f'A{row}:C{row}')
            summary_cell = ws[f'A{row}']
            summary_cell.value = summary_line
            summary_cell.font = Font(size=11)
            summary_cell.border = border
            summary_cell.alignment = left_alignment
            ws.row_dimensions[row].height = 20
            row += 1
        
        # Save the workbook
        output_file = "TEST_SUMMARY_REPORT.xlsx"
        wb.save(output_file)
        
        print(f"\n✅ Excel report saved to: {output_file}")
        
        print("\n" + "="*80)
        print("✅ TEST SUMMARY REPORT GENERATED SUCCESSFULLY!")
        print("="*80)
        print("\n📊 METRICS SUMMARY:")
        print(f"   • Total Test Cases         : {total_test_cases}")
        print(f"   • Total Executed Tests     : {total_executed}")
        print(f"   • Passed Tests ✅          : {total_passed}")
        print(f"   • Failed Tests ❌          : {total_failed}")
        print(f"   • Skipped Tests ⏭️        : {total_test_cases - total_executed}")
        print(f"\n   • Pass Rate %              : {pass_rate:.2f}%")
        print(f"   • Fail Rate %              : {fail_rate:.2f}%")
        print(f"   • Execution Rate %         : {execution_rate:.2f}%")
        print("\n" + "="*80)
        print(f"\n📁 Generated File: {output_file}")
        print("\n")
        
        return True
        
    except FileNotFoundError:
        print(f"❌ Error: File '{excel_file}' not found!")
        print("   Please ensure the Excel file is in the same directory.")
        return False
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    excel_file = 'MSME-Base-Test-Cases-V1.xlsx'
    analyze_excel_and_generate_xlsx_report(excel_file)
